#!/usr/bin/env python3

"""Converts all excel files in the CWD to similarly named CSV files."""

import os
import csv
import openpyxl

for file in os.listdir('.'):
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)

        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            name_splice = file[: -5]
            csv_file = open(name_splice + '_' + sheet_name + '.csv', 'w', newline='')
            csv_writer = csv.writer(csv_file)

            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row_num, column=col_num).value)

                for row in row_data:
                    csv_writer.writerow(row)
            csv_file.close()
