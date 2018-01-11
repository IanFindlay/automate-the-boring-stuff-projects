#!/usr/bin/env python3

"""Takes each column of a spreadsheet and saves it to a seperate text file."""

import re
import openpyxl

print('Enter the absolute path of the spreasheet you wish to split:')
file_path = input()

path_regex = re.compile(r'(.*/)(.*)(\.xlsx)$')
path_split = path_regex.search(file_path)
path = path_split.group(1)
name = path_split.group(2)

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

col_num = 1
for column in range(1, sheet.max_column + 1):

    col_data = []
    for cell in range(1, sheet.max_row + 1):
        if sheet.cell(row=cell, column=col_num).value != None:
            col_data.append(sheet.cell(row=cell, column=col_num).value)

    file = open('/' + path + 'col-' + str(col_num) + '-' + name + '.txt', 'w')
    for line in col_data:
        file.write(line + '\n')
    file.close()

    col_num += 1
