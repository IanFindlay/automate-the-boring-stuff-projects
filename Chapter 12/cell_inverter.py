#!/usr/bin/env python3

"""Invert the rows and columns of a spreadsheet."""

import re
import openpyxl

print('Enter the absolute path to the spreadsheet file:')
file_path = input()

path_regex = re.compile(r'(.*/)(.*)(\.xlsx)$')
path_split = path_regex.search(file_path)
path = path_split.group(1)
name = path_split.group(2)

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Make nested list of spreadsheet data (row by row)
rows = []
max_row = sheet.max_row
max_column = sheet.max_column
for row in range(1, max_row + 1):
    data = []
    for cell in range(1, max_column + 1):
        cell_value = sheet.cell(row=row, column=cell).value
        data.append(cell_value)
    rows.append(data)

# Open new spreadsheet and populate with above data but inverted
wb = openpyxl.Workbook()
sheet = wb.active

column_num = 1
for row in rows:
    row_num = 1
    for cell in row:
        sheet.cell(row=row_num, column=column_num).value = cell
        row_num += 1
    column_num += 1

wb.save(path + name + '(inverted).xlsx')
print('Spreadsheet data inverted and saved to ' + name + '(inverted).xlsx.')
