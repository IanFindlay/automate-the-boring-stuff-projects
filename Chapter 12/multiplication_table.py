#!/usr/bin/env python3

"""Creates a N x N multiplication table in a spreadsheet.

Usage: multiplication_table.py <N> - Creates N X N table
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

number = int(sys.argv[1])
cells = number + 1

wb = openpyxl.Workbook()
sheet = wb.active

make_bold = Font(bold=True)

while cells > 1:
    # Write outer cells values and make them bold
    sheet.cell(row=cells, column=1).value = cells - 1
    sheet.cell(row=1, column=cells).value = cells - 1

    sheet.cell(row=cells, column=1).font = make_bold
    sheet.cell(row=1, column=cells).font = make_bold

    cells -= 1

# Populate the table innards with correct formula
col_length = number + 1

count = 0
while count < number:
    col_letter = get_column_letter(sheet.max_column - count)

    while col_length > 1:
        sheet[col_letter + str(col_length)] = ('=SUM(' + col_letter + '1*A'
                                               + str(col_length) + ')')

        col_length -= 1

    col_length = number + 1
    count += 1

wb.save('multi_table.xlsx')
